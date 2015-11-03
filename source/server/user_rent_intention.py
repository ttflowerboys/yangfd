# coding: utf-8
from bson.objectid import ObjectId
from datetime import datetime
from datetime import timedelta
from app import f_app
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

f_app.common.memcache_server = ["172.20.101.98:11211"]
f_app.common.mongo_server = "172.20.101.98"


def get_data_directly(user, part, deep=None):
    if user is None:
        return
    user_part = user.get(part, None)
    if (user_part is not None) and (deep is not None):
        return user.get(part).get(deep, None)
    if f_app.util.batch_iterable(user_part):
        user_part = '/'.join(user_part)
    return user_part


def get_data_complex(user, target, condition, element):

    '''this func make dict provide get_data_enum to use.
    with user's id and 'condition' to search in the database 'target'
    then gether element in search result, make a new dict return
    '''

    dic = {}
    user_id = user.get("id", None)
    target_database = getattr(f_app, target)
    condition.update({"$or": [{"user_id": ObjectId(user_id)},
                              {"creator_user_id": ObjectId(user_id)}]})
    # select_item = target_database.get(target_database.search(condition, per_page=10))
    select_item = target_database.get(target_database.search(condition, per_page=-1))
    element_list = []
    for ticket in select_item:
        element_list.append(ticket.get(element, None))
    dic.update({element: element_list})
    return dic


def get_data_enum(user, enum_name):
    if user is None:
        return
    single = user.get(enum_name, None)
    value_list = []
    if f_app.util.batch_iterable(single):
        for true_single in single:
            if true_single is None:
                continue
            enum_id = true_single.get("id", None)
            value = enum_type_list[enum_name].get(enum_id, None)
            value_list.append(value)
    elif single is not None:
        if single.get("id", None):
            enum_id = str(single.get("id", None))
            value = enum_type_list[enum_name].get(enum_id, None)
            #print enum_id
            #print enum_type_list[enum_name]
            if value is not None:
                value_list.append(value)
    if not f_app.util.batch_iterable(value_list):
        value_list = [value_list]
    value_set = set(value_list)
    value_list = list(value_set)
    return '/'.join(value_list)


def get_diff_color(fill, total):
    base_color = 0x999999
    color = 0x0
    for index in range(total):
        s = 0x222222
        color = base_color + s*index
        if total <= 4:
            color_t = '00'+"%x" % color
            fill.append(PatternFill(fill_type='solid', start_color=color_t, end_color=color_t))


def format_fit(sheet):
    simsun_font = Font(name="SimSun")
    header_fill = PatternFill(fill_type='solid', start_color='00dddddd', end_color='00dddddd')
    alignment_fit = Alignment(shrink_to_fit=True)
    for cell in sheet.rows[0]:
        cell.fill = header_fill
    for row in sheet.rows:
        for cell in row:
            cell.font = simsun_font
            cell.alignment = alignment_fit
    for num, col in enumerate(sheet.columns):
        lenmax = 0
        col_set = set()
        for line, cell in enumerate(col):
            if line != 0:
                col_set.add(cell.value)
            lencur = 0
            if isinstance(cell.value, int) or isinstance(cell.value, datetime):
                lencur = len(str(cell.value).encode("GBK"))
            elif cell.value is not None:
                lencur = len(cell.value.encode("GBK", "replace"))
            if lencur > lenmax:
                lenmax = lencur
        if num > 90:
            sheet.column_dimensions['A'+chr(num-26)]
            print "col "+'A'+chr(num-26)+" fit."
        else:
            sheet.column_dimensions[chr(num+65)].width = lenmax*0.86
            print "col "+chr(num+65)+" fit."
        cell_fill = []
        if 1 < len(col_set) <= 4:
            get_diff_color(cell_fill, len(col_set))
            for color_index, fill_index in enumerate(list(col_set)):
                for line, cell in enumerate(col):
                    if line == 0:
                        continue
                    if fill_index == cell.value and len(cell.value):
                        cell.fill = cell_fill[color_index]


def get_all_rent_intention():
    params = {"type": "rent_intention"}
    return f_app.ticket.output(f_app.ticket.search(params, per_page=-1))
    # return f_app.i18n.process_i18n(f_app.ticket.output(f_app.ticket.search(params, per_page=-1)))


def get_referer(time):
    diff_time = timedelta(milliseconds=100)
    flag = 0
    record = referer_result[0]
    for num, single in enumerate(referer_result):
        rst_time = single.get("time", None)
        if rst_time is None:
            continue
        if rst_time - diff_time < time < rst_time + diff_time:
            if flag:
                print "bingo. "+str(rst_time)
            flag = 1
            record = single
    if flag:
        return record.get("referer", '')
    return '找不到'


enum_type = ["rent_type"]
enum_type_list = {}
for enum_singlt_type in enum_type:
    print "enum type " + enum_singlt_type + " loading."
    enum_list_subdic = {}
    for enumitem in f_app.i18n.process_i18n(f_app.enum.get_all(enum_singlt_type)):
        enum_list_subdic.update({enumitem["id"]: enumitem["value"]})
    enum_type_list.update({enum_singlt_type: enum_list_subdic})
    print "enum type " + enum_singlt_type + " done."

referer_result = f_app.log.output(f_app.log.search({"route": "/api/1/rent_intention_ticket/add"}, per_page=-1))
header = ["状态", "标题", "客户", "联系方式", "提交时间", "起始日期", "出租需求", "预算上限", "预算下限", "period", "出租位置", "备注",
          "样房东有无匹配搭配", "referer", "打电话了？", "有接到？", "房子租到了么？", "通过样房东",
          "如果不是通过样房东，那么是通过哪里什么样的房源？有没有交中介费", "对平台体验的想法及反馈",
          "在找房子中用户最疼的点有哪些？", "备注"]


wb = Workbook()
ws = wb.active

ws.append(header)
print "loading..."
for number, ticket in enumerate(get_all_rent_intention()):
    print 'ticket.' + str(number) + ' loading.'
    ticket = f_app.i18n.process_i18n(ticket)
    print 'ticket.' + str(number) + ' i18n process complete.'
    period_start = get_data_directly(ticket, "rent_available_time")
    period_end = get_data_directly(ticket, "rent_deadline_time")
    if period_end is None or period_start is None:
        time = "不明"
    else:
        period = period_end - period_start
        if period.days >= 365:
            time = "longer than 12 months"
        elif 365 > period.days >= 180:
            time = "6 - 12 months"
        elif 180 >= period.days > 90:
            time = "3 - 6 months"
        elif period.days <= 30:
            time = "less than 1 month"
    city = ticket.get("city", {})
    country = ticket.get("country", {})
    maponics_neighborhood = ticket.get("maponics_neighborhood", {})
    match = []
    if "partial_match" in ticket.get("tags", []):
        match.append("部分满足")
    if "perfect_match" in ticket.get("tags", []):
        match.append("完全满足")

    ws.append(["已提交" if (get_data_directly(ticket, "status") == "new") else "已出租",
               get_data_directly(ticket, "title"),
               get_data_directly(ticket, "nickname"),
               get_data_directly(ticket, "phone"),
               get_data_directly(ticket, "time"),
               get_data_directly(ticket, "rent_available_time"),
               get_data_enum(ticket, "rent_type"),
               get_data_directly(ticket, "rent_budget_max", "value"),
               get_data_directly(ticket, "rent_budget_min", "value"),
               time,
               ' '.join([country.get("code", ''),
                         city.get("name", ''),
                         maponics_neighborhood.get("name", ''),
                         ticket.get("address", ''),
                         ticket.get("zipcode_index", '')]),
               get_data_directly(ticket, "description"),
               '/'.join(match),
               get_referer(get_data_directly(ticket, "time"))
               ])
    print 'ticket.' + str(number) + ' done.'
format_fit(ws)
wb.save("user_rent_intention.xlsx")
